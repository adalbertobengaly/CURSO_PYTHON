# type: ignore
# openpyxl - Ler e alterar dados de planilhas do Workbook
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook, load_workbook
# from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOT_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo do Excel
workbook: Workbook = load_workbook(WORKBOOT_PATH)
# worksheet: Worksheet = workbook.active

# Nome para a planilha
sheet_name = 'Minha planilha'

# Selecionou a planinha
worksheet: Worksheet = workbook[sheet_name]

for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 23)
    print()

# worksheet['B3'].value = 14

workbook.save(WORKBOOT_PATH)
